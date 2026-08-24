#!/usr/bin/env python3
"""置信度校准分析：各置信度档"声称把握" vs "实际命中率"对照 + Brier 分数。
用法: python3 scripts/calibration.py
门槛：MIN_N=20 个"已记录置信度且已方向结算"样本前只输出进度（幂等安全，可每日跑）；
达标后生成 reports/calibration_最新.md 并打印结论。"""
import os, datetime
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOG = f"{ROOT}/log.xlsx"
MIN_N = 20
BJ = datetime.timezone(datetime.timedelta(hours=8))

wb = load_workbook(LOG, data_only=True)
ws = wb["预测日志"]
header = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(header) if h}
if "置信度(%)" not in idx:
    print("log.xlsx 无 置信度(%) 列，先按 README 补列"); raise SystemExit(1)

samples = []  # (日期, 置信度, 命中1/0)
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[idx["目标日期"]] is None: continue
    conf, dok = r[idx["置信度(%)"]], r[idx["方向对错"]]
    if conf is None or dok not in ("√", "×"): continue
    d = r[idx["目标日期"]]
    d = d.isoformat()[:10] if hasattr(d, "isoformat") else str(d)[:10]
    samples.append((d, float(conf), 1 if dok == "√" else 0))

n = len(samples)
print(f"置信度校准：已结算且带置信度的样本 {n}/{MIN_N}")
if n < MIN_N:
    left = MIN_N - n
    print(f"还差 {left} 个样本（约 {left} 个交易日后达标），达标前只记进度、不出长报告。")
    raise SystemExit(0)

# ---- 分档对照 ----
def bucket(c):
    if c < 55: return "<55"
    if c < 65: return "55-64"
    if c < 75: return "65-74"
    return "≥75"

groups = {}
for d, c, o in samples:
    groups.setdefault(bucket(c), []).append((c, o))

rows, total_gap = [], 0.0
for b in ["<55", "55-64", "65-74", "≥75"]:
    g = groups.get(b, [])
    if not g: continue
    stated = sum(c for c, _ in g) / len(g)
    hit = sum(o for _, o in g) / len(g) * 100
    gap = hit - stated  # 正=保守低估，负=过度自信
    total_gap += gap * len(g)
    rows.append((b, len(g), stated, hit, gap))

brier = sum((c / 100 - o) ** 2 for _, c, o in samples) / n
overall_gap = total_gap / n
verdict = ("校准良好（整体偏差 |{:.1f}pp| ≤ 5）".format(overall_gap) if abs(overall_gap) <= 5
           else ("过度自信（整体高估 {:.1f}pp，建议晨间推理下调置信度）".format(-overall_gap) if overall_gap < 0
                 else "偏保守（整体低估 {:.1f}pp，可适当上调）".format(overall_gap)))

now = datetime.datetime.now(BJ)
lines = [
    "# 置信度校准报告", "",
    f"**生成时间**：{now.strftime('%Y-%m-%d %H:%M')}（北京时间）｜ **样本量**：{n} 个已结算预测",
    f"**整体结论**：{verdict}｜ **Brier 分数**：{brier:.4f}（0=完美，0.25=瞎猜，越小越好）", "",
    "| 置信度档 | 样本数 | 平均声称 | 实际命中率 | 偏差(pp) | 解读 |",
    "|---|---|---|---|---|---|",
]
for b, cnt, stated, hit, gap in rows:
    read = "校准良好" if abs(gap) <= 5 else ("高估·过度自信" if gap < 0 else "低估·偏保守")
    lines.append(f"| {b}% | {cnt} | {stated:.1f}% | {hit:.1f}% | {gap:+.1f} | {read} |")
lines += [
    "", "**读法**：某档『实际命中率』显著低于『平均声称』（偏差 < -5pp）说明模型在该档过度自信，"
    "反之显著高于则是保守。据此可在晨间推理时对应上调/下调该档置信度。", "",
    f"**口径**：方向对错按 log.xlsx M 列（最终方向 vs 实际）；置信度取报告结论表的方向置信度。"
    f"样本 {n} 个，每档样本 <5 时该档结论仅供参考。", "",
]

os.makedirs(f"{ROOT}/reports", exist_ok=True)
path = f"{ROOT}/reports/calibration_{now.strftime('%Y-%m-%d')}.md"
with open(path, "w", encoding="utf-8") as f: f.write("\n".join(lines))
print(f"校准报告已生成: {path}")
print(verdict + f"｜Brier {brier:.4f}")
for b, cnt, stated, hit, gap in rows:
    print(f"  {b}%: n={cnt} 声称{stated:.0f}% 实际{hit:.0f}% 偏差{gap:+.1f}pp")

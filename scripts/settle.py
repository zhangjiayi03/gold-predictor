#!/usr/bin/env python3
"""结算昨日预测：取实时金价填入 log.xlsx 中目标日期=昨天(北京时间)的行，公式自动计分。
用法: python3 settle.py   （无待结算行时自动跳过，幂等安全）"""
import json, os, datetime, subprocess, urllib.request
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # 仓库根（git 克隆内运行，勿用绝对路径）
LOG = f"{ROOT}/log.xlsx"
RECALC = "/data/user/builtin/work/default/skills/xlsx/scripts/recalc.py"
BJ = datetime.timezone(datetime.timedelta(hours=8))
TODAY = datetime.datetime.now(BJ).date().isoformat()  # 北京时间日期（沙盒时钟为UTC，直接today()会差8小时）
TARGET = (datetime.datetime.now(BJ).date() - datetime.timedelta(days=1)).isoformat()  # 待结算行：目标日期=昨天（被预测的交易日已走完）

# 1. 实时金价（gold-api 挂了退出码 2，由任务层决定是否 AgentEarth 兜底）
try:
    req = urllib.request.Request("https://api.gold-api.com/price/XAU",
                                 headers={"User-Agent": "Mozilla/5.0"})
    price = json.loads(urllib.request.urlopen(req, timeout=25).read())["price"]
except Exception as e:
    print(f"GOLD_API_FAIL: {e}")
    raise SystemExit(2)
print(f"实时金价: ${price:.2f}")

# 2. 定位目标日期=今天且实际价为空的行
wb = load_workbook(LOG)          # 保留公式，勿用 data_only
ws = wb["预测日志"]
header = [c.value for c in ws[1]]
col_date = header.index("目标日期") + 1
col_actual = next(i+1 for i, h in enumerate(header) if h and str(h).startswith("次日实际价"))

def as_date(v):
    if hasattr(v, "isoformat"): return v.isoformat()[:10]
    return str(v).strip()[:10] if v else None

hit = False
for row in ws.iter_rows(min_row=2):
    if as_date(row[col_date-1].value) != TARGET: continue
    if row[col_actual-1].value is not None:
        print(f"目标日期 {TARGET} 已结算过，跳过（幂等）"); break
    row[col_actual-1] = round(price, 2)
    hit = True
    print(f"已结算: 目标日期={TARGET}（昨日交易日）实际价=${price:.2f}（J-N 公式自动计分）")
    break

if not hit:
    print("无待结算行（无目标日期=昨天且实际价为空的行）")
else:
    wb.save(LOG)
    try:  # 3. 重算公式，确保零错误
        r = subprocess.run(["python3", RECALC, LOG], capture_output=True, text=True, timeout=120)
        print("recalc:", (r.stdout or r.stderr).strip()[:200])
    except Exception as e:
        print(f"recalc 跳过: {e}（请手动跑 {RECALC}）")

#!/usr/bin/env python3
# 黄金预测仪表盘生成器：读取 log.xlsx + data/*.json + predictions/*.md，重写固定页面
import json, glob, os, re, datetime, urllib.request
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # 仓库根（git 克隆内运行，勿用绝对路径）
DASH = f"{ROOT}/dashboard"

def load_json(p):
    with open(p, encoding="utf-8") as f: return json.load(f)

# ---- 1. 价格序列 ----
pb = load_json(f"{ROOT}/data/price_backfill.json")
series = pb["series"]
base = pb["baseline"]
changes = []
for i in range(1, len(series)):
    chg = (series[i]["close"] - series[i-1]["close"]) / series[i-1]["close"] * 100
    changes.append([series[i]["date"], round(chg, 2)])

# ---- 2. 最新采集 ----
collects = sorted(glob.glob(f"{ROOT}/data/*_collect.json"))
collect = load_json(collects[-1]) if collects else {"news": {}, "indicators": {}}
live = collect.get("indicators", {}).get("实时金价XAU") or {}
live_price = live.get("price")
live_time = live.get("updatedAt", "")

# ---- 3. 日志行 ----
wb = load_workbook(f"{ROOT}/log.xlsx", data_only=False)
ws = wb["预测日志"]
rows, header = [], [c.value for c in ws[1]]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[0] is None: continue
    rows.append(dict(zip(["date","gen","base","fdir","fband","mdir","mband","manual","actual","chg","adir","aband","dok","bok","regime","drivers","note"], r)))

settled = []
for r in rows:
    if r["actual"] is not None and r["base"]:
        chg = (r["actual"] - r["base"]) / r["base"]
        adir = "涨" if chg > 0 else ("跌" if chg < 0 else "平")
        aband = "小" if abs(chg) <= 0.003 else ("中" if abs(chg) <= 0.01 else "大")
        r["pct"] = round(chg * 100, 2)
        r["adir"], r["aband"] = adir, aband
        r["dok"] = "√" if r["fdir"] == adir else "×"
        r["bok"] = "√" if r["fband"] == aband else "×"
        settled.append(r)

dir_ok = sum(1 for r in settled if r["dok"] == "√")
band_ok = sum(1 for r in settled if r["bok"] == "√")
cum_dir, cum_band, xs = [], [], []
for i, r in enumerate(settled, 1):
    xs.append(r["date"])
    cum_dir.append(round(sum(1 for x in settled[:i] if x["dok"] == "√") / i * 100, 1))
    cum_band.append(round(sum(1 for x in settled[:i] if x["bok"] == "√") / i * 100, 1))

# ---- 4. 最新预测报告解析 ----
preds = sorted(glob.glob(f"{ROOT}/predictions/*.md"))
bulls, bears, pm_review = [], [], ""
if preds:
    md = open(preds[-1], encoding="utf-8").read()
    mb = re.search(r"### 利多.*?\n(.*?)### 利空", md, re.S)
    ms = re.search(r"### 利空.*?\n(.*?)(?:###|$)", md, re.S)
    def items(block):
        if not block: return []
        return [re.sub(r"^\d+\.\s*", "", l.strip().lstrip("- "))[:90] for l in block.group(1).split("\n") if l.strip().startswith(("1.", "2.", "3.", "4.", "5.", "-"))][:5]
    bulls, bears = items(mb), items(ms)
    pm_all = re.findall(r"### 午后复核\s*\n(.*?)(?:\n### |\n## |\Z)", md, re.S)
    if pm_all: pm_review = re.sub(r"\s+", " ", pm_all[-1]).strip()[:280]  # 取最后一个：同日多次复核时永远显示最新
    plain_m = re.search(r"##\s*💬?\s*大白话解读\s*\n(.*?)(?=\n## |\Z)", md, re.S)
    plain = ""
    if plain_m:
        lines = [re.sub(r"\*\*", "", l.strip()) for l in plain_m.group(1).split("\n") if l.strip() and l.strip() != "---"]
        plain = "｜".join([l for l in lines if not l.startswith("**今天为什么")][:6])

latest = rows[-1] if rows else {}
conf_m = re.search(r"\*\*方向\*\* \| \*\*(\w+)\*\* \| (\d+)%", open(preds[-1], encoding="utf-8").read()) if preds else None
conf = conf_m.group(2) if conf_m else "—"

# ---- 4.5 美元兑人民币汇率（open.er-api.com 免Key，frankfurter 备用） ----
BJ = datetime.timezone(datetime.timedelta(hours=8))
usdcny = None
for u in ["https://open.er-api.com/v6/latest/USD", "https://api.frankfurter.dev/v1/latest?base=USD&symbols=CNY"]:
    try:
        req = urllib.request.Request(u, headers={"User-Agent": "Mozilla/5.0"})
        j = json.loads(urllib.request.urlopen(req, timeout=20).read())
        usdcny = j.get("rates", {}).get("CNY") or j.get("cny")
        if usdcny: break
    except Exception:
        continue
if not usdcny:
    try:
        usdcny = load_json(f"{ROOT}/data/usdcny_last.json")["rate"]  # 兜底用上次
    except Exception:
        usdcny = 6.74
os.makedirs(f"{ROOT}/data", exist_ok=True)
with open(f"{ROOT}/data/usdcny_last.json", "w") as f:
    json.dump({"rate": usdcny, "time": datetime.datetime.now(BJ).strftime("%Y-%m-%d %H:%M")}, f, ensure_ascii=False)

# ---- 5. 新闻摘要 ----
news_groups = []
for cat, items in (collect.get("news") or {}).items():
    arr = [{"t": x.get("title","")[:80], "u": x.get("url","")} for x in items[:4] if isinstance(x, dict) and "error" not in x]
    if arr: news_groups.append({"cat": cat, "items": arr})

# ---- 6. 组装数据 ----
data = {
    "updated": datetime.datetime.now(BJ).strftime("%Y-%m-%d %H:%M"),
    "nextTarget": latest.get("date", ""),
    "prediction": {
        "dir": latest.get("fdir", "—"), "band": latest.get("fband", "—"), "conf": conf,
        "base": latest.get("base"), "drivers": latest.get("drivers", ""),
        "regime": latest.get("regime") or base.get("regime_20d", "—"),
        "bulls": bulls, "bears": bears, "pm": pm_review, "plain": plain
    },
    "live": {"price": live_price, "time": live_time, "usdcny": usdcny},
    "priceSeries": series,
    "dailyChanges": changes,
    "baseline": base,
    "stats": {"total": len(rows), "settled": len(settled), "dirOK": dir_ok, "bandOK": band_ok},
    "accuracy": {"dates": xs, "dir": cum_dir, "band": cum_band},
    "history": [
        {"date": r["date"], "base": r["base"], "fdir": r["fdir"], "fband": r["fband"],
         "mdir": r["mdir"], "mband": r["mband"], "manual": r["manual"],
         "actual": r["actual"], "pct": r.get("pct"), "adir": r.get("adir"), "aband": r.get("aband"),
         "dok": r.get("dok",""), "bok": r.get("bok",""), "regime": r["regime"], "drivers": r["drivers"]}
        for r in reversed(rows)
    ],
    "news": news_groups
}
os.makedirs(f"{DASH}/assets", exist_ok=True)
with open(f"{DASH}/assets/data.js", "w", encoding="utf-8") as f:
    f.write("window.DASH_DATA = " + json.dumps(data, ensure_ascii=False, indent=1) + ";")
print(f"data.js written: {len(rows)} rows, {len(settled)} settled, live={live_price}")
